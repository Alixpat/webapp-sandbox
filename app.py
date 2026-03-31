import io
import os
import uuid
from datetime import datetime, timezone

from flask import (
    Flask,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from flask_login import (
    LoginManager,
    UserMixin,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL", "postgresql://hello:hello@localhost:5432/hellodb"
)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-key-change-me")
app.config["UPLOAD_FOLDER"] = os.path.join(os.path.dirname(__file__), "uploads")
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16 MB

os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = "Veuillez vous connecter."

# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------

class User(UserMixin, db.Model):
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    role = db.Column(db.String(20), nullable=False, default="user")  # user | supervisor
    projects = db.relationship("Project", backref="owner", lazy=True)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    @property
    def is_supervisor(self):
        return self.role == "supervisor"


class Project(db.Model):
    __tablename__ = "projects"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, default="")
    created_at = db.Column(db.DateTime, nullable=False, default=lambda: datetime.now(timezone.utc))
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    budget_needs = db.relationship("BudgetNeedExpression", backref="project", lazy=True, cascade="all, delete-orphan")

    @property
    def total_amount(self):
        return sum(bn.total_amount for bn in self.budget_needs)

    @property
    def total_mco(self):
        return sum(bn.amount_mco for bn in self.budget_needs)

    @property
    def total_investment(self):
        return sum(bn.amount_investment for bn in self.budget_needs)


class BudgetNeedExpression(db.Model):
    __tablename__ = "budget_need_expressions"
    id = db.Column(db.Integer, primary_key=True)
    label = db.Column(db.String(200), nullable=False)
    amount_mco = db.Column(db.Float, nullable=False, default=0)
    amount_investment = db.Column(db.Float, nullable=False, default=0)
    justification = db.Column(db.Text, default="")
    created_at = db.Column(db.DateTime, nullable=False, default=lambda: datetime.now(timezone.utc))
    project_id = db.Column(db.Integer, db.ForeignKey("projects.id"), nullable=False)
    attachments = db.relationship("Attachment", backref="budget_need", lazy=True, cascade="all, delete-orphan")

    @property
    def total_amount(self):
        return self.amount_mco + self.amount_investment


class Attachment(db.Model):
    __tablename__ = "attachments"
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(256), nullable=False)
    original_name = db.Column(db.String(256), nullable=False)
    uploaded_at = db.Column(db.DateTime, nullable=False, default=lambda: datetime.now(timezone.utc))
    budget_need_id = db.Column(db.Integer, db.ForeignKey("budget_need_expressions.id"), nullable=False)


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


# ---------------------------------------------------------------------------
# Seed data
# ---------------------------------------------------------------------------

def seed_data():
    if User.query.first():
        return

    # Users
    alice = User(name="Alice Dupont", email="alice@example.com", role="user")
    alice.set_password("alice123")
    bob = User(name="Bob Martin", email="bob@example.com", role="user")
    bob.set_password("bob123")
    admin = User(name="Sophie Leroy", email="admin@example.com", role="supervisor")
    admin.set_password("admin123")
    db.session.add_all([alice, bob, admin])
    db.session.flush()

    # Projects for Alice
    p1 = Project(name="Refonte site web", description="Modernisation du site web corporate avec nouveau design et migration vers le cloud.", user_id=alice.id)
    p2 = Project(name="Formation Data Science", description="Programme de formation en data science pour l'équipe technique.", user_id=alice.id)
    db.session.add_all([p1, p2])
    db.session.flush()

    # Budget needs for project 1
    db.session.add_all([
        BudgetNeedExpression(label="Design UX/UI", amount_mco=0, amount_investment=15000, justification="Prestation agence de design pour maquettes et prototypes interactifs.", project_id=p1.id),
        BudgetNeedExpression(label="Développement front-end", amount_mco=0, amount_investment=25000, justification="Développement React.js par un prestataire externe sur 3 mois.", project_id=p1.id),
        BudgetNeedExpression(label="Hébergement cloud", amount_mco=4800, amount_investment=0, justification="Abonnement AWS pour 12 mois (EC2 + RDS + S3).", project_id=p1.id),
    ])

    # Budget needs for project 2
    db.session.add_all([
        BudgetNeedExpression(label="Licence plateforme e-learning", amount_mco=3000, amount_investment=0, justification="Abonnement annuel DataCamp Team pour 10 utilisateurs.", project_id=p2.id),
        BudgetNeedExpression(label="Intervenant externe", amount_mco=0, amount_investment=8000, justification="2 jours de formation en présentiel par un expert Machine Learning.", project_id=p2.id),
    ])

    # Project for Bob
    p3 = Project(name="Migration ERP", description="Migration de l'ERP legacy vers une solution SaaS moderne.", user_id=bob.id)
    db.session.add(p3)
    db.session.flush()

    db.session.add_all([
        BudgetNeedExpression(label="Audit technique", amount_mco=0, amount_investment=12000, justification="Audit complet de l'existant par un cabinet de conseil.", project_id=p3.id),
        BudgetNeedExpression(label="Licence ERP SaaS", amount_mco=36000, amount_investment=0, justification="Abonnement annuel Odoo Enterprise pour 25 utilisateurs.", project_id=p3.id),
        BudgetNeedExpression(label="Conduite du changement", amount_mco=0, amount_investment=5000, justification="Accompagnement des équipes et documentation utilisateur.", project_id=p3.id),
    ])

    db.session.commit()


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    if current_user.is_authenticated:
        if current_user.is_supervisor:
            return redirect(url_for("supervisor_dashboard"))
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(password):
            login_user(user)
            next_page = request.args.get("next")
            return redirect(next_page or url_for("index"))
        flash("Email ou mot de passe incorrect.", "danger")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# --- User dashboard ---

@app.route("/dashboard")
@login_required
def dashboard():
    projects = Project.query.filter_by(user_id=current_user.id).order_by(Project.created_at.desc()).all()
    return render_template("dashboard.html", projects=projects)


@app.route("/projects/new", methods=["GET", "POST"])
@login_required
def project_new():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        if not name:
            flash("Le nom du projet est obligatoire.", "danger")
        else:
            project = Project(name=name, description=description, user_id=current_user.id)
            db.session.add(project)
            db.session.commit()
            flash("Projet créé avec succès.", "success")
            return redirect(url_for("project_detail", project_id=project.id))
    return render_template("project_new.html")


@app.route("/projects/<int:project_id>")
@login_required
def project_detail(project_id):
    project = db.session.get(Project, project_id)
    if not project:
        flash("Projet introuvable.", "danger")
        return redirect(url_for("dashboard"))
    if not current_user.is_supervisor and project.user_id != current_user.id:
        flash("Accès non autorisé.", "danger")
        return redirect(url_for("dashboard"))
    budget_needs = BudgetNeedExpression.query.filter_by(project_id=project.id).order_by(BudgetNeedExpression.created_at.desc()).all()
    return render_template("project_detail.html", project=project, budget_needs=budget_needs)


@app.route("/projects/<int:project_id>/budget-needs/new", methods=["GET", "POST"])
@login_required
def budget_need_new(project_id):
    project = db.session.get(Project, project_id)
    if not project or project.user_id != current_user.id:
        flash("Accès non autorisé.", "danger")
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        label = request.form.get("label", "").strip()
        amount_mco_str = request.form.get("amount_mco", "").strip() or "0"
        amount_investment_str = request.form.get("amount_investment", "").strip() or "0"
        justification = request.form.get("justification", "").strip()
        if not label:
            flash("Le libellé est obligatoire.", "danger")
        else:
            try:
                amount_mco = float(amount_mco_str)
                amount_investment = float(amount_investment_str)
            except ValueError:
                flash("Montant invalide.", "danger")
                return render_template("budget_need_new.html", project=project)
            if amount_mco == 0 and amount_investment == 0:
                flash("Au moins un des deux montants (MCO ou Investissement) doit être renseigné.", "danger")
                return render_template("budget_need_new.html", project=project)
            bn = BudgetNeedExpression(label=label, amount_mco=amount_mco, amount_investment=amount_investment, justification=justification, project_id=project.id)
            db.session.add(bn)
            db.session.flush()

            files = request.files.getlist("attachments")
            for f in files:
                if f and f.filename:
                    original = secure_filename(f.filename)
                    ext = os.path.splitext(original)[1]
                    stored = f"{uuid.uuid4().hex}{ext}"
                    f.save(os.path.join(app.config["UPLOAD_FOLDER"], stored))
                    att = Attachment(filename=stored, original_name=original, budget_need_id=bn.id)
                    db.session.add(att)

            db.session.commit()
            flash("Expression de besoin budgétaire ajoutée.", "success")
            return redirect(url_for("project_detail", project_id=project.id))
    return render_template("budget_need_new.html", project=project)


@app.route("/uploads/<filename>")
@login_required
def uploaded_file(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


# --- Supervisor dashboard ---

@app.route("/supervisor")
@login_required
def supervisor_dashboard():
    if not current_user.is_supervisor:
        flash("Accès réservé aux superviseurs.", "danger")
        return redirect(url_for("dashboard"))
    projects = Project.query.order_by(Project.created_at.desc()).all()
    total = sum(p.total_amount for p in projects)
    return render_template("supervisor.html", projects=projects, total=total)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _style_header(ws, col_count):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="000091", end_color="000091", fill_type="solid")
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


@app.route("/projects/<int:project_id>/export")
@login_required
def export_project(project_id):
    project = db.session.get(Project, project_id)
    if not project:
        flash("Projet introuvable.", "danger")
        return redirect(url_for("dashboard"))
    if not current_user.is_supervisor and project.user_id != current_user.id:
        flash("Accès non autorisé.", "danger")
        return redirect(url_for("dashboard"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Expressions de besoin"

    headers = ["Libellé", "MCO (€)", "Investissement (€)", "Total (€)", "Justification", "Pièces jointes", "Date"]
    ws.append(headers)
    _style_header(ws, len(headers))

    budget_needs = BudgetNeedExpression.query.filter_by(project_id=project.id).order_by(BudgetNeedExpression.created_at.desc()).all()
    for bn in budget_needs:
        attachments = ", ".join(att.original_name for att in bn.attachments)
        ws.append([bn.label, bn.amount_mco, bn.amount_investment, bn.total_amount, bn.justification, attachments, bn.created_at.strftime("%d/%m/%Y")])

    # Total row
    total_row = len(budget_needs) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=project.total_mco).font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=project.total_investment).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=project.total_amount).font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"budget_{project.name.replace(' ', '_')}.xlsx"
    return send_file(buf, download_name=filename, as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/supervisor/export")
@login_required
def export_supervisor():
    if not current_user.is_supervisor:
        flash("Accès réservé aux superviseurs.", "danger")
        return redirect(url_for("dashboard"))

    projects = Project.query.order_by(Project.created_at.desc()).all()

    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Synthèse"
    headers = ["Projet", "Propriétaire", "Nb expressions", "MCO (€)", "Investissement (€)", "Budget total (€)", "Date de création"]
    ws_summary.append(headers)
    _style_header(ws_summary, len(headers))

    for p in projects:
        ws_summary.append([p.name, p.owner.name, len(p.budget_needs), p.total_mco, p.total_investment, p.total_amount, p.created_at.strftime("%d/%m/%Y")])

    total_row = len(projects) + 2
    ws_summary.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws_summary.cell(row=total_row, column=4, value=sum(p.total_mco for p in projects)).font = Font(bold=True)
    ws_summary.cell(row=total_row, column=5, value=sum(p.total_investment for p in projects)).font = Font(bold=True)
    ws_summary.cell(row=total_row, column=6, value=sum(p.total_amount for p in projects)).font = Font(bold=True)

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 20
    ws_summary.column_dimensions["C"].width = 16
    ws_summary.column_dimensions["D"].width = 15
    ws_summary.column_dimensions["E"].width = 18
    ws_summary.column_dimensions["F"].width = 18
    ws_summary.column_dimensions["G"].width = 15

    # Sheet 2: All budget need expressions
    ws_detail = wb.create_sheet("Détail expressions de besoin")
    detail_headers = ["Projet", "Propriétaire", "Libellé", "MCO (€)", "Investissement (€)", "Total (€)", "Justification", "Date"]
    ws_detail.append(detail_headers)
    _style_header(ws_detail, len(detail_headers))

    for p in projects:
        for bn in p.budget_needs:
            ws_detail.append([p.name, p.owner.name, bn.label, bn.amount_mco, bn.amount_investment, bn.total_amount, bn.justification, bn.created_at.strftime("%d/%m/%Y")])

    ws_detail.column_dimensions["A"].width = 25
    ws_detail.column_dimensions["B"].width = 20
    ws_detail.column_dimensions["C"].width = 30
    ws_detail.column_dimensions["D"].width = 15
    ws_detail.column_dimensions["E"].width = 18
    ws_detail.column_dimensions["F"].width = 15
    ws_detail.column_dimensions["G"].width = 50
    ws_detail.column_dimensions["H"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name="budget_global.xlsx", as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------------------
# Initialisation
# ---------------------------------------------------------------------------

def _migrate_old_schema():
    """Migrate from old budget_lines schema to budget_need_expressions."""
    from sqlalchemy import inspect as sa_inspect, text

    inspector = sa_inspect(db.engine)
    existing_tables = inspector.get_table_names()

    if "budget_lines" not in existing_tables:
        return  # No migration needed

    with db.engine.begin() as conn:
        # 1. Create the new table
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS budget_need_expressions (
                id SERIAL PRIMARY KEY,
                label VARCHAR(200) NOT NULL,
                amount_mco FLOAT NOT NULL DEFAULT 0,
                amount_investment FLOAT NOT NULL DEFAULT 0,
                justification TEXT DEFAULT '',
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                project_id INTEGER NOT NULL REFERENCES projects(id)
            )
        """))

        # 2. Migrate data (put old amount into amount_investment)
        conn.execute(text("""
            INSERT INTO budget_need_expressions (id, label, amount_mco, amount_investment, justification, created_at, project_id)
            SELECT id, label, 0, amount, justification, created_at, project_id
            FROM budget_lines
            WHERE id NOT IN (SELECT id FROM budget_need_expressions)
        """))

        # 3. Update attachments FK column
        att_columns = [c["name"] for c in inspector.get_columns("attachments")]
        if "budget_line_id" in att_columns and "budget_need_id" not in att_columns:
            conn.execute(text(
                "ALTER TABLE attachments RENAME COLUMN budget_line_id TO budget_need_id"
            ))
            # Update FK constraint (PostgreSQL)
            try:
                conn.execute(text(
                    "ALTER TABLE attachments DROP CONSTRAINT IF EXISTS attachments_budget_line_id_fkey"
                ))
                conn.execute(text(
                    "ALTER TABLE attachments ADD CONSTRAINT attachments_budget_need_id_fkey "
                    "FOREIGN KEY (budget_need_id) REFERENCES budget_need_expressions(id)"
                ))
            except Exception:
                pass  # SQLite doesn't support DROP/ADD CONSTRAINT

        # 4. Sync the serial sequence so new inserts get correct IDs (PostgreSQL)
        try:
            conn.execute(text(
                "SELECT setval('budget_need_expressions_id_seq', "
                "(SELECT COALESCE(MAX(id), 0) FROM budget_need_expressions))"
            ))
        except Exception:
            pass  # SQLite has no sequences

        # 5. Drop old table
        conn.execute(text("DROP TABLE IF EXISTS budget_lines CASCADE"))


with app.app_context():
    _migrate_old_schema()
    db.create_all()
    seed_data()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
